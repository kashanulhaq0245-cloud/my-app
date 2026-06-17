import os
import openpyxl
import time
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from app.config import settings

# Ensure reports directory exists inside uploads
reports_dir = os.path.join(settings.UPLOAD_DIR, "reports")
os.makedirs(reports_dir, exist_ok=True)

def generate_excel_report(vehicles) -> str:
    """
    Generate an Excel sheet of vehicle logs.
    Returns the absolute path to the generated Excel file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ANPR Vehicle Log Report"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "AI ANPR Smart Gate System - Vehicle Log Report"
    title_cell.font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Gray
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = ["Log ID", "Vehicle Number", "Entry Time", "Exit Time", "Date", "Status", "Image Link"]
    ws.append([]) # Empty row
    ws.append(headers)
    ws.row_dimensions[3].height = 25
    
    header_fill = PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    for col_num in range(1, 8):
        cell = ws.cell(row=3, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    # Data rows
    for index, v in enumerate(vehicles, start=4):
        entry_t = v.entry_time.strftime("%Y-%m-%d %H:%M:%S") if v.entry_time else "-"
        exit_t = v.exit_time.strftime("%Y-%m-%d %H:%M:%S") if v.exit_time else "-"
        date_str = v.date.strftime("%Y-%m-%d") if v.date else "-"
        img_url = f"http://localhost:8000{v.photo_path}" if v.photo_path else "No Photo"
        
        row_data = [v.id, v.vehicle_number, entry_t, exit_t, date_str, v.status, img_url]
        ws.append(row_data)
        ws.row_dimensions[index].height = 20
        
        # Style row cells
        status_color = "DEF7EC" if v.status == "Inside" else "FDE8E8" # Light green vs light red
        status_text_color = "03543F" if v.status == "Inside" else "9B1C1C"
        
        for col_num in range(1, 8):
            cell = ws.cell(row=index, column=col_num)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            
            # Highlight status column
            if col_num == 6:
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                cell.font = Font(name="Arial", size=10, bold=True, color=status_text_color)
                
            # Make image link clickable
            if col_num == 7 and v.photo_path:
                cell.hyperlink = img_url
                cell.font = Font(name="Arial", size=10, color="2563EB", underline="single")
                
    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    filepath = os.path.join(reports_dir, f"report_{int(time.time())}.xlsx")
    wb.save(filepath)
    return filepath

def generate_pdf_report(vehicles) -> str:
    """
    Generate a styled PDF report of vehicle logs using ReportLab.
    Returns the absolute path to the generated PDF.
    """
    filepath = os.path.join(reports_dir, f"report_{int(time.time())}.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#1F2937'),
        spaceAfter=10,
        alignment=1 # Centered
    )
    
    meta_style = ParagraphStyle(
        'MetaText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#4B5563'),
        spaceAfter=15,
        alignment=1
    )
    
    story = []
    
    # Add Header
    story.append(Paragraph("AI ANPR Smart Gate System - Vehicle Log Report", title_style))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total Records: {len(vehicles)}", meta_style))
    story.append(Spacer(1, 10))
    
    # Table headers
    data = [["ID", "Vehicle Number", "Entry Time", "Exit Time", "Date", "Status"]]
    
    # Populate data rows
    for v in vehicles:
        entry_t = v.entry_time.strftime("%H:%M:%S") if v.entry_time else "-"
        exit_t = v.exit_time.strftime("%H:%M:%S") if v.exit_time else "-"
        date_str = v.date.strftime("%Y-%m-%d") if v.date else "-"
        
        data.append([
            str(v.id),
            v.vehicle_number,
            entry_t,
            exit_t,
            date_str,
            v.status
        ])
        
    # Create Table
    col_widths = [40, 120, 100, 100, 90, 80]
    table = Table(data, colWidths=col_widths)
    
    # Styling Table
    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F2937')), # Dark gray header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ])
    
    # Highlight individual cell rows based on status
    for i in range(1, len(data)):
        status = data[i][5]
        if status == "Inside":
            t_style.add('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#03543F'))
            t_style.add('BACKGROUND', (5, i), (5, i), colors.HexColor('#DEF7EC'))
        else:
            t_style.add('TEXTCOLOR', (5, i), (5, i), colors.HexColor('#9B1C1C'))
            t_style.add('BACKGROUND', (5, i), (5, i), colors.HexColor('#FDE8E8'))
            
    table.setStyle(t_style)
    story.append(table)
    
    # Build Document
    doc.build(story)
    return filepath
